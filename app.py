import openpyxl
from flask import Flask, render_template, request
from flask_caching import Cache

app = Flask(__name__)
# cache = Cache(app, config={'CACHE_TYPE': 'simple'})

# Define the route for the main page
@app.route('/')
# @cache.cached(timeout=60)

def index():
    # Read the Excel file and select the worksheet
    wb = openpyxl.load_workbook('Test.xlsx')
    ws = wb.active

    # Get the headers for the Excel data
    headers = [cell.value for cell in ws[1]]

    data = []
    for row in ws.iter_rows(min_row=2, max_row=6):
            print(row[0].value)
            data.append([cell.value for cell in row])
            # print(data)

    # Render the HTML template with the headers and empty data
    return render_template('index.html', headers=headers, data=data)

# Define the route for the search functionality
@app.route('/search', methods=['GET', 'POST'])
# @cache.cached(timeout=60)

def search():
    # Read the Excel file and select the worksheet
    wb = openpyxl.load_workbook('Test.xlsx')
    ws = wb.active

    # Get the headers for the Excel data
    headers = [cell.value for cell in ws[1]]

    # Get the search term from the form data
    search_term = request.form['search']

    # Filter the data based on the search term
    data = []
    for row in ws.iter_rows(min_row=2):
        if (search_term.lower() in row[2].value.lower()) or (search_term.lower() in row[4].value.lower()):
            data.append([cell.value for cell in row])

    # Render the HTML template with the headers and filtered data
    return render_template('index.html', headers=headers, data=data)

if __name__ == '__main__':
    app.run(debug=True)
